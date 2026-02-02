from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Count
import json
from datetime import datetime, date, timedelta
from .forms import UploadFileForm
from .services import procesar_archivo_activos, asignar_turnos_automatico
from .models import Turno, Responsable, Equipo, ConfiguracionCronograma, Feriado

def index(request):
    return redirect('ver_cronograma')

def upload_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                procesar_archivo_activos(request.FILES['archivo'])
                if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
                    return JsonResponse({'status': 'ok', 'message': "Archivo procesado con éxito."})
                messages.success(request, "Archivo procesado con éxito.")
                return redirect('ver_cronograma')
            except Exception as e:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
                    return JsonResponse({'status': 'error', 'message': f"Error al procesar el archivo: {e}"}, status=400)
                messages.error(request, f"Error al procesar el archivo: {e}")
    return redirect('ver_cronograma')

def api_get_datos(request):
    """
    Endpoint para obtener todos los datos necesarios para re-renderizar el cronograma.
    """
    config = ConfiguracionCronograma.objects.last()
    turnos = Turno.objects.select_related('responsable').prefetch_related('equipos').all()
    feriados = Feriado.objects.all()
    
    turnos_data = []
    for t in turnos:
        turnos_data.append({
            'id': t.id,
            'responsable': t.responsable.nombre,
            'fecha': t.fecha.strftime('%Y-%m-%d') if t.fecha else '',
            'hora': t.hora.strftime('%H:%M') if t.hora else '',
            'estado': t.estado,
            'equipos': [{
                'id': e.id,
                'marca': e.marca,
                'modelo': e.modelo,
                'codigo': e.codigo,
                'descripcion': e.descripcion,
                'atendido': e.atendido
            } for e in t.equipos.all()]
        })

    data = {
        'turnos': turnos_data,
        'feriados': [f.fecha.strftime('%Y-%m-%d') for f in feriados],
        'config': {
            'inicio': config.fecha_inicio.strftime('%Y-%m-%d') if config and config.fecha_inicio else '',
            'fin': config.fecha_fin.strftime('%Y-%m-%d') if config and config.fecha_fin else '',
        }
    }
    return JsonResponse(data)

def get_day_shifts(request, date):
    """
    Endpoint para obtener todos los turnos de un día específico.
    """
    turnos = Turno.objects.filter(fecha=date).select_related('responsable').prefetch_related('equipos').all()
    
    turnos_data = []
    for t in turnos:
        turnos_data.append({
            'id': t.id,
            'responsable': t.responsable.nombre,
            'fecha': t.fecha.strftime('%Y-%m-%d') if t.fecha else '',
            'hora': t.hora.strftime('%H:%M') if t.hora else '',
            'estado': t.estado,
            'equipos': [{
                'id': e.id,
                'marca': e.marca,
                'modelo': e.modelo,
                'codigo': e.codigo,
                'descripcion': e.descripcion,
                'atendido': e.atendido
            } for e in t.equipos.all()]
        })
    
    return JsonResponse({'turnos': turnos_data, 'fecha': date})

def ver_cronograma(request):
    config = ConfiguracionCronograma.objects.last()
    turnos = Turno.objects.select_related('responsable').prefetch_related('equipos').all()
    feriados = Feriado.objects.all()
    
    # Datos para el sidebar
    responsables_stats = Responsable.objects.annotate(num_equipos=Count('equipos'))
    
    context = {
        'config': config,
        'turnos': turnos,
        'feriados': feriados,
        'responsables': responsables_stats,
        'form': UploadFileForm(),
    }
    return render(request, 'core/cronograma.html', context)

@require_POST
def guardar_configuracion(request):
    data = request.POST
    try:
        config, created = ConfiguracionCronograma.objects.get_or_create(id=1)
        
        # Asignar valores directamente (los campos ahora aceptan null)
        if data.get('fecha_inicio'):
            config.fecha_inicio = data.get('fecha_inicio')
        if data.get('fecha_fin'):
            config.fecha_fin = data.get('fecha_fin')
        if data.get('hora_inicio'):
            config.hora_inicio = data.get('hora_inicio')
        if data.get('hora_fin'):
            config.hora_fin = data.get('hora_fin')
        if data.get('hora_almuerzo'):
            config.hora_almuerzo = data.get('hora_almuerzo')
        if data.get('duracion_turno'):
            config.duracion_turno = int(data.get('duracion_turno'))
        if data.get('duracion_almuerzo'):
            config.duracion_almuerzo = int(data.get('duracion_almuerzo'))
        
        if data.get('modo_exclusion'):
            config.modo_exclusion = data.get('modo_exclusion')
        
        f_inicio = data.get('fecha_inicio')
        f_fin = data.get('fecha_fin')

        if not f_inicio or not f_fin:
            return JsonResponse({
                'status': 'error',
                'message': 'Las fechas de inicio y fin son obligatorias.'
            }, status=400)

        if f_inicio > f_fin:
            return JsonResponse({
                'status': 'error', 
                'message': 'La fecha de inicio no puede ser posterior a la fecha de fin.'
            }, status=400)

        # Validación de rango de horas
        h_inicio = data.get('hora_inicio')
        h_fin = data.get('hora_fin')
        if h_inicio and h_fin:
            if h_inicio >= h_fin:
                return JsonResponse({
                    'status': 'error',
                    'message': 'La hora de inicio debe ser anterior a la hora de fin.'
                }, status=400)

        config.save()
        return JsonResponse({'status': 'ok', 'message': 'Configuración guardada correctamente'})
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error en guardar_configuracion: {error_detail}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def generar_cronograma_view(request):
    try:
        count, result = asignar_turnos_automatico()
        if count > 0:
            return JsonResponse({'status': 'ok', 'message': result})
        else:
            # result puede ser un string o un dict con detalles del error
            data = result if isinstance(result, dict) else {'message': result}
            return JsonResponse({'status': 'error', 'data': data}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error', 
            'data': {'message': f'Error interno del servidor: {str(e)}'}
        }, status=500)

@require_POST
def actualizar_turno(request, turno_id):
    turno = get_object_or_404(Turno, id=turno_id)
    data = json.loads(request.body)
    
    if 'fecha' in data:
        turno.fecha = data['fecha']
    if 'hora' in data:
        turno.hora = data['hora']
    if 'estado' in data:
        turno.estado = data['estado']
        
    turno.save()
    return JsonResponse({'status': 'ok'})

@require_POST
def intercambiar_turnos(request):
    """
    Intercambia la fecha y hora de dos turnos específicos.
    """
    try:
        data = json.loads(request.body)
        turno_a_id = data.get('turno_a_id')
        turno_b_id = data.get('turno_b_id')
        
        if not turno_a_id or not turno_b_id:
            return JsonResponse({'status': 'error', 'message': 'IDs de turno no proporcionados'}, status=400)
            
        from django.db import transaction
        with transaction.atomic():
            turno_a = get_object_or_404(Turno, id=turno_a_id)
            turno_b = get_object_or_404(Turno, id=turno_b_id)
            
            # Intercambiar fecha y hora
            fecha_temp, hora_temp = turno_a.fecha, turno_a.hora
            turno_a.fecha, turno_a.hora = turno_b.fecha, turno_b.hora
            turno_b.fecha, turno_b.hora = fecha_temp, hora_temp
            
            turno_a.save()
            turno_b.save()
            
        return JsonResponse({'status': 'ok', 'message': 'Turnos intercambiados correctamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def toggle_completado(request, turno_id):
    turno = get_object_or_404(Turno, id=turno_id)
    if turno.estado == 'completado':
        turno.estado = 'asignado'
    else:
        turno.estado = 'completado'
        # Si se marca como completado, marcar todos sus equipos como atendidos
        turno.equipos.update(atendido=True)
        
    turno.save()
    
    # Devolver el estado de los equipos para actualizar el panel lateral
    equipos_data = list(turno.equipos.values('id', 'atendido'))
    
    return JsonResponse({
        'status': 'ok', 
        'nuevo_estado': turno.estado,
        'equipos': equipos_data
    })

@require_POST
def toggle_equipo_atendido(request, equipo_id):
    equipo = get_object_or_404(Equipo, id=equipo_id)
    equipo.atendido = not equipo.atendido
    equipo.save()
    
    # Calcular estadísticas del TURNO
    turno = equipo.turno
    if not turno:
        return JsonResponse({'status': 'error', 'message': 'Equipo no vinculado a un turno'}, status=400)
        
    total_equipos = turno.equipos.count()
    equipos_atendidos = turno.equipos.filter(atendido=True).count()
    
    # Sincronizar con el estado del Turno
    if equipos_atendidos == total_equipos and total_equipos > 0:
        turno.estado = 'completado'
    elif equipos_atendidos > 0:
        turno.estado = 'en_proceso'
    else:
        turno.estado = 'asignado'
    turno.save()
    
    return JsonResponse({
        'status': 'ok',
        'atendido': equipo.atendido,
        'equipos_atendidos': equipos_atendidos,
        'total_equipos': total_equipos,
        'turno_estado': turno.estado
    })

@require_POST
def add_feriado(request):
    data = json.loads(request.body)
    fecha = data.get('fecha')
    if fecha:
        Feriado.objects.get_or_create(fecha=fecha)
        feriados = list(Feriado.objects.values_list('fecha', flat=True))
        return JsonResponse({'status': 'ok', 'feriados': [f.strftime('%Y-%m-%d') for f in feriados]})
    return JsonResponse({'status': 'error', 'message': 'Fecha inválida'}, status=400)

@require_POST
def remove_feriado(request):
    data = json.loads(request.body)
    fecha = data.get('fecha')
    if fecha:
        Feriado.objects.filter(fecha=fecha).delete()
        feriados = list(Feriado.objects.values_list('fecha', flat=True))
        return JsonResponse({'status': 'ok', 'feriados': [f.strftime('%Y-%m-%d') for f in feriados]})
    return JsonResponse({'status': 'error', 'message': 'Fecha inválida'}, status=400)

def exportar_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    # 1. Obtener Datos (Incluir turnos asignados, en proceso y completados)
    turnos = Turno.objects.filter(estado__in=['asignado', 'en_proceso', 'completado']).select_related('responsable').prefetch_related('equipos')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma de Mantenimiento"

    # 2. Definir Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde claro
    success_font = Font(color="006100", bold=True)
    process_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo claro
    process_font = Font(color="9C5700", bold=True)
    
    info_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Azul claro
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # 3. Encabezado del Reporte
    ws.merge_cells('A1:F1')
    ws['A1'] = "SISTEMA TECHSCHEDULER - CRONOGRAMA DE MANTENIMIENTO PREVENTIVO"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws['A1'].alignment = center_align

    # 4. Resumen Ejecutivo
    total = turnos.count()
    completados = turnos.filter(estado='completado').count()
    pendientes = total - completados
    
    ws['A3'] = "RESUMEN DE EJECUCIÓN"
    ws['A3'].font = Font(bold=True, size=10, underline="single")
    
    ws['A4'] = "Total Responsables:"; ws['B4'] = total
    ws['A5'] = "Mantenimientos Listos:"; ws['B5'] = completados; ws['B5'].font = success_font
    ws['A6'] = "Mantenimientos Pendientes:"; ws['B6'] = pendientes; ws['B6'].font = Font(color="9C0006", bold=True)
    
    ws['D4'] = "Fecha Reporte:"; ws['E4'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('D4:E4')

    # 5. Tabla de Datos
    headers = ['FECHA', 'HORA', 'RESPONSABLE', 'EQUIPOS', 'ESTADO', 'DETALLE TÉCNICO (MARCA, MODELO, ID)']
    
    start_row = 8
    for i, h in enumerate(headers):
        cell = ws.cell(row=start_row, column=i+1)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 6. Llenado de Datos
    current_row = start_row + 1
    for t in turnos:
        # Fecha y Hora
        c_fecha = ws.cell(row=current_row, column=1, value=t.fecha)
        c_fecha.alignment = center_align
        c_fecha.number_format = 'DD/MM/YYYY'
        
        ws.cell(row=current_row, column=2, value=t.hora.strftime("%H:%M") if t.hora else "--:--").alignment = center_align
        
        # Responsable
        ws.cell(row=current_row, column=3, value=t.responsable.nombre).alignment = left_align
        
        # Equipos (conteo del turno)
        eq_all = t.equipos.all()
        eq_count = eq_all.count()
        atendidos = eq_all.filter(atendido=True).count()
        ws.cell(row=current_row, column=4, value=f"{atendidos}/{eq_count}").alignment = center_align
        
        # Estado con Color
        if t.estado == 'completado':
            status_text = "LISTO"
        elif t.estado == 'en_proceso' or (atendidos > 0 and atendidos < eq_count):
            status_text = "EN PROCESO"
        else:
            status_text = "PENDIENTE"
            
        status_cell = ws.cell(row=current_row, column=5, value=status_text)
        status_cell.alignment = center_align
        
        if status_text == 'LISTO':
            status_cell.fill = success_fill
            status_cell.font = success_font
        elif status_text == 'EN PROCESO':
            # Amarillo más potente
            status_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            status_cell.font = Font(color="9C5700", bold=True)
        else:
            status_cell.fill = info_fill
            status_cell.font = Font(color="0070C0", bold=True)
        
        # Detalle granular por equipo (Incluye ID GOBIERNO)
        detalles_lista = []
        for eq in eq_all:
            lbl = "[LISTO]" if eq.atendido else "[PENDIENTE]"
            id_str = f"({eq.codigo})" if eq.codigo else "(S/N)"
            detalles_lista.append(f"{lbl} {eq.marca} {eq.modelo} {id_str}")
            
        detalles_full = " | ".join(detalles_lista)
        ws.cell(row=current_row, column=6, value=detalles_full).alignment = left_align
        ws.cell(row=current_row, column=6).font = Font(size=8)
        
        # Bordes
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = border
            
        current_row += 1

    # 7. Ajuste de Estética Final
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 85

    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Cronograma_Mantenimiento_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response

@require_POST
def reset_database(request):
    """
    Limpia todos los datos de la base de datos para empezar de cero.
    """
    try:
        from django.db import transaction
        with transaction.atomic():
            Turno.objects.all().delete()
            Equipo.objects.all().delete()
            Responsable.objects.all().delete()
            Feriado.objects.all().delete()
            ConfiguracionCronograma.objects.all().delete()
        return JsonResponse({'status': 'ok', 'message': 'Sistema reiniciado correctamente.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def crear_turno_manual(request):
    """
    Crea un nuevo turno manualmente con un nuevo responsable y equipos.
    """
    try:
        data = json.loads(request.body)
        print("DEBUG: Payload received for crear_turno_manual:", data)
        nombre = data.get('nombre')
        email = data.get('email')
        fecha_str = data.get('fecha')
        hora_str = data.get('hora')
        equipos = data.get('equipos', [])

        if not nombre or not fecha_str or not hora_str:
            return JsonResponse({'status': 'error', 'message': 'Nombre, Fecha y Hora son obligatorios.'}, status=400)

        # Convert strings to objects
        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            hora_obj = datetime.strptime(hora_str, '%H:%M').time()
        except ValueError as ve:
            return JsonResponse({'status': 'error', 'message': f'Formato de fecha u hora inválido: {ve}'}, status=400)

        from django.db import transaction
        import traceback
        with transaction.atomic():
            # 1. Verificar si el nombre ya existe
            existing_responsable = Responsable.objects.filter(nombre=nombre).first()
            
            # Si existe y no se ha confirmado explícitamente el duplicado, avisar al usuario
            force_create = data.get('force_create', False)
            
            if existing_responsable and not force_create:
                # Buscar si ya tiene un turno HOY o en el futuro para advertir mejor
                # pero ahora permitimos tener varios históricos. Solo advertimos del nombre duplicado.
                return JsonResponse({
                    'status': 'duplicate',
                    'message': f'Ya existe un perfil grabado para "{nombre}". ¿Desea crear un nuevo turno independiente para esta persona?',
                    'existing_responsable': {
                        'nombre': existing_responsable.nombre,
                        'email': existing_responsable.email or 'Sin correo'
                    }
                })
            
            # 2. Obtener o Crear Responsable
            responsable, created = Responsable.objects.get_or_create(nombre=nombre)
            
            # Ya NO eliminamos el turno existente (force_create simplemente procede a crear uno NUEVO)
            
            if email:
                responsable.email = email
                responsable.save()
            
            # 3. Crear Turno (NUEVO e independiente)
            turno = Turno.objects.create(
                responsable=responsable,
                fecha=fecha_obj,
                hora=hora_obj,
                estado='asignado'
            )
            
            # 4. Crear Equipos vinculados al TURNO
            for eq in equipos:
                Equipo.objects.create(
                    turno=turno,
                    responsable=responsable,
                    codigo=eq.get('codigo'),
                    marca=eq.get('marca'),
                    modelo=eq.get('modelo'),
                    descripcion=eq.get('descripcion')
                )
            
            # 5. Sincronizar cola de notificaciones
            try:
                from notifications.services import NotificationService
                NotificationService.sincronizar_cola()
            except:
                pass

        return JsonResponse({
            'status': 'ok', 
            'message': 'Turno creado correctamente.',
            'turno_id': turno.id
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
